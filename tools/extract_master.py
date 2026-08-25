# -*- coding: utf-8 -*-
"""엑셀(QSC·MS 평가표.xlsx) → data/master.json 추출.

엑셀 시트가 문항·심각도의 원본. 시트 수정 후 이 스크립트를 다시 실행하면
앱 마스터 데이터가 갱신된다. 실행: python tools/extract_master.py

v3.7 구조(2026-08-13):
 - 74문항(행 10~83), 판정(○△X) 폐지 — '개선 필요 건수'만 입력
 - 심각도 = I열 코드(S1=★★ 즉시위해 / S2=★ 중대운영 / 공란=일반)
 - 절대 감점제: 일반 1건 −1 / ★ 문항당 −8·추가 −2·합계 상한 −45
   / ★★ 문항당 −12·추가 −4·합계 상한 −48. QSC = MAX(0, 100 − 일반 감점)
 - 대분류 가중치 폐지(H열 태그는 영역 참고표 전용)
 - 쇼퍼: 문항 행은 D열 유효성에서 읽는다(상수 아님). 예/아니오 · 5점 척도 두 갈래
 - 등급: 우수 93점 이상(2026-08-12 결정)
"""
import sys, io, json, re, warnings
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')

SRC = r'C:\Users\glow-pc-017\Desktop\Ai\1. QSC\QSC·MS 평가표.xlsx'
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'data' / 'master.json'

# ★버전을 손으로 적지 않는다★ — 적어 두면 엑셀만 고치고 이 줄을 안 고쳤을 때
#   master.json이 옛 날짜를 그대로 달고 나간다. 화면 세 곳이 이 값을 '평가표 … 기준'으로
#   보여 주므로, 그 눈금이 거짓말을 하면 배포가 됐는지 사람이 알 길이 없다.
import datetime, hashlib
_raw = Path(SRC).read_bytes()
XLSX_VERSION = datetime.datetime.fromtimestamp(Path(SRC).stat().st_mtime).strftime('%Y-%m-%d')
XLSX_SHA = hashlib.sha256(_raw).hexdigest()[:12]     # 사람에게는 안 보인다. 기계가 대조용으로 쓴다.

wb = openpyxl.load_workbook(SRC)

# ★2026-08-25: 시트 두 장을 한 장으로 합쳤다★ (사용자 요청)
#   전   「QSC 평가표」 · 「미스터리쇼퍼 평가표」 · 「채점기준」 · 「예비항목」  네 장
#   후   「QSC·MS 평가표」 한 장 — QSC 는 A~I, 미스터리쇼퍼는 ★S열부터★
#   (채점기준·예비항목은 이 스크립트가 읽지 않던 시트라 없애도 영향이 없다)
#
#   그래서 쇼퍼 쪽은 같은 시트를 보되 ★열을 MSC 만큼 밀어★ 읽는다.
#   행 번호는 그대로다(나란히 붙였다). 옮길 때 엑셀이 수식의 열 참조도 같이 밀어 두었다
#   (D11 → V11 식으로. 93칸 전부 대조 확인).
SHEET = 'QSC·MS 평가표'
MSC = 18                 # 쇼퍼 A열(1) → S열(19). 즉 +18
ws = wb[SHEET]

# ★행 범위를 박아 두지 않는다★ — A열 병합이 곧 대분류 경계다.
#   상수로 두면 엑셀에서 행을 하나 넣고 하나 빼는 순간, 74 검사는 통과하면서
#   문항이 옆 대분류로 넘어간다(쇼퍼 쪽에서 같은 부류의 사고가 실제로 났다).
GROUPS = sorted((m.min_row, m.max_row) for m in ws.merged_cells.ranges
                if m.min_col == 1 and 10 <= m.min_row <= 83)
assert len(GROUPS) == 6, 'QSC 대분류가 %d개 — A열 병합을 확인하십시오' % len(GROUPS)

qsc_groups = []
item_no = 0
sev_count = {'S1': 0, 'S2': 0}
for g0, g1 in GROUPS:
    gname = ws[f'A{g0}'].value
    items = []
    for r in range(g0, g1 + 1):
        item_no += 1
        text = str(ws[f'C{r}'].value or '').strip()
        m = re.match(r'^([A-F]-\d{2}[ab]?)\.\s*(.*)$', text)
        code = m.group(1) if m else None
        if m:
            text = m.group(2)
        sev = str(ws[f'I{r}'].value or '').strip()      # S1 / S2 / ''
        if sev not in ('S1', 'S2'):
            sev = ''
        if sev:
            sev_count[sev] += 1
        items.append({
            'no': item_no, 'code': code, 'row': r, 'text': text,
            'severity': sev,                 # '' | 'S2'(★) | 'S1'(★★)
            'critical': bool(sev),           # 구버전 호환 필드
        })
    name = str(gname).strip().replace('\n', ' ') if gname else ''
    qsc_groups.append({'name': name, 'count': g1 - g0 + 1, 'items': items})

assert item_no == 74, f'문항 수 {item_no} — 74가 아님. GROUPS 범위 확인 필요'
assert sev_count == {'S1': 4, 'S2': 12}, f'심각도 분포 {sev_count} — ★★4/★12가 아님. I열 확인 필요'

sh = ws                  # 같은 시트다. 열만 MSC 만큼 밀어서 읽는다.

# ── 쇼퍼 문항 범위 ─────────────────────────────────────────────────
# ★행 번호를 박아 두지 않는다★ — D열 데이터 유효성이 곧 문항 범위다.
#   '예,아니오,NA' = 관찰 문항 / '1,2,3,4,5' = 만족도(5점 척도) 문항.
#   2026-08-18에 13번 카테고리가 4문항 → 2문항으로 줄었는데, 종전 코드는
#   LIKERT_ROWS = range(41,51)과 assert q_no == 40을 상수로 들고 있어 그대로 깨졌다.
#   문항이 늘거나 줄 때 사람이 이 파일까지 같이 고쳐야 하는 구조를 없앤다.
def _rows_for(choices):
    out = set()
    for dv in sh.data_validations.dataValidation:
        if str(dv.formula1 or '').strip().strip('"') != choices:
            continue
        for rng in dv.sqref.ranges:
            if rng.min_col <= MSC + 4 <= rng.max_col:    # 쇼퍼 D열(=V열)을 덮는 규칙만
                out.update(range(rng.min_row, rng.max_row + 1))
    return out


YN_ROWS = _rows_for('예,아니오,NA')
LIKERT_ROWS = _rows_for('1,2,3,4,5')
Q_ROWS = YN_ROWS | LIKERT_ROWS
assert Q_ROWS, '쇼퍼 문항 행을 못 찾음 — D열 유효성 규칙을 확인하십시오'
assert not (YN_ROWS & LIKERT_ROWS), f'두 척도에 겹쳐 걸린 행 {sorted(YN_ROWS & LIKERT_ROWS)}'
Q_MIN, Q_MAX = min(Q_ROWS), max(Q_ROWS)

# 카테고리 이름은 쇼퍼 A열(=S열)에 있다 — 병합이면 시작 행에만, 한 줄짜리면 그 행에 있다
cat_at = {}
for m in sh.merged_cells.ranges:
    if m.min_col == MSC + 1 and Q_MIN <= m.min_row <= Q_MAX:
        cat_at[m.min_row] = sh.cell(m.min_row, MSC + 1).value
for r in sorted(Q_ROWS):
    v = sh.cell(r, MSC + 1).value
    if v not in (None, ''):
        cat_at.setdefault(r, v)

shopper_cats = []
q_no = 0
cur = None
for r in sorted(Q_ROWS):
    if r in cat_at:
        cur = {'name': str(cat_at[r]).strip().replace('\n', ' '), 'questions': []}
        shopper_cats.append(cur)
    text = sh.cell(r, MSC + 3).value                     # 쇼퍼 C열 = U열
    if not text:
        continue
    assert cur is not None, f'{r}행 문항 위에 카테고리 이름이 없습니다'
    q_no += 1
    cur['questions'].append({
        'no': q_no, 'row': r, 'text': str(text).strip(),
        'scale': 'likert' if r in LIKERT_ROWS else 'yn',
    })
shopper_cats = [c for c in shopper_cats if c['questions']]
n_likert = len(LIKERT_ROWS)
assert q_no == len(Q_ROWS), f'문항 {q_no}개 / 유효성 행 {len(Q_ROWS)}개 — 질문 칸이 빈 행이 있습니다'

# 비고 문구도 행을 찾아 읽는다 (종전에는 C66으로 박혀 있었다)
_note_row = next((r for r in range(Q_MAX + 1, Q_MAX + 30)
                  if str(sh.cell(r, MSC + 1).value or '').strip() == '비고'), None)
shopper_note = sh.cell(_note_row, MSC + 3).value if _note_row else None

# 매장 목록: 구글 통합시트에서 **직접** 읽는다 (2026-08-20).
#
# ★ 왜 바꿨는가
#   전에는 로컬 폴더의 `[감사총무팀_QSC] 통합시트.xlsx` 를 읽었다. 그런데 그 파일은
#   **참고용으로 받아둔 사본**이지 원본이 아니다. 원본은 항상 구글 스프레드시트다.
#   사본이 낡으면 옛 매장 목록이 조용히 배포됐다 — 실제로 '창창족발 (창창 창신)' 이
#   구글에서 '창창 창신' 으로 바뀐 뒤 10일간 앱에 옛 이름이 나갔다.
#   매장 수(26)는 그대로여서 아래 개수 검사에도 걸리지 않았다.
#
# ★ 왜 CSV 인가 (xlsx 내려받기 대신)
#   gviz CSV 는 파일을 만들지 않고 값만 받는다(17KB vs 679KB). 무엇보다
#   **숨김 행을 구글이 알아서 걸러서** 보내준다 — 숨김 = 폐점·검사 제외라는 규칙이
#   시트 쪽에 있으니, 그 판정을 여기서 다시 하지 않는 편이 어긋날 여지가 없다.
#   (2026-08-20 대조: CSV 26곳 = xlsx 표시행 26곳 = 당시 앱 26곳, 완전 일치)
#
# ★ 못 받으면 멈춘다
#   낡은 목록이 조용히 나가는 것보다 배포가 막히는 편이 낫다. 지금까지의 사고가
#   전부 '틀린 게 조용히 나가는 것' 이었다.
import csv as _csv, io as _io, urllib.parse as _up, urllib.request as _ur

SHEET_URL = 'https://docs.google.com/spreadsheets/d/1FSlYCWz4zNToV1hkPUQJvv4RYCBfc1ecf5l46YFD4Qk/edit?gid=448253562#gid=448253562'
_SHEET_TAB = '데이터'
_STORE_COL = '매장명'

def _fetch_stores():
    m = re.search(r'/spreadsheets/d/([A-Za-z0-9_-]+)', SHEET_URL)
    if not m:
        raise SystemExit('★중단★ SHEET_URL 에서 시트 ID를 찾지 못했습니다.')
    url = ('https://docs.google.com/spreadsheets/d/%s/gviz/tq?tqx=out:csv&sheet=%s'
           % (m.group(1), _up.quote(_SHEET_TAB)))
    try:
        req = _ur.Request(url, headers={'User-Agent': 'qsc-extract'})
        raw = _ur.urlopen(req, timeout=30).read().decode('utf-8')
    except Exception as e:
        raise SystemExit('★중단★ 구글 통합시트를 읽지 못했습니다 (%s). '
                         '인터넷 연결과, 시트 공유가 "링크가 있는 사람은 볼 수 있음" 인지 '
                         '확인해 주세요.' % e)
    rows = list(_csv.reader(_io.StringIO(raw)))
    if not rows:
        raise SystemExit('★중단★ 구글 통합시트가 비어 있습니다.')
    hdr = [(h or '').strip() for h in rows[0]]
    if _STORE_COL not in hdr:
        raise SystemExit("★중단★ '%s' 열을 찾지 못했습니다. 머리글: %s"
                         % (_STORE_COL, ', '.join(h for h in hdr[:8] if h)))
    ci = hdr.index(_STORE_COL)
    out = []
    for r in rows[1:]:
        if len(r) > ci and (r[ci] or '').strip():
            out.append(r[ci].strip())
    return out

stores = _fetch_stores()
print('매장 %d 개 (구글 통합시트 직독 · 숨김 행은 구글이 제외)' % len(stores))
assert len(stores) >= 20, ('★중단★ 매장이 %d곳뿐입니다 — 구글 통합시트 「%s」 탭의 '
                           "'%s' 열을 확인하십시오." % (len(stores), _SHEET_TAB, _STORE_COL))

master = {
    'version': XLSX_VERSION,          # 화면에 '평가표 … 기준'으로 뜬다
    'source_sha': XLSX_SHA,           # 엑셀이 바뀌면 반드시 바뀐다 — 배포 점검이 이걸 본다
    'source': 'QSC·MS 평가표.xlsx',
    'stores': stores,
    # 절대 감점제 파라미터
    # ⚠2026-08-25에 엑셀 '채점기준' 시트를 없앴다(사용자 요청). 그래서 이 숫자들을
    #   대조할 시트가 이제 없다 — ★여기가 사실상의 원본★이다. 고칠 때는 아래 세 곳을
    #   반드시 함께 본다: 이 파일 · js/scoring.js · backend/Code.gs.
    #   옛 '채점기준' 시트가 필요하면 _보관/작업기록/백업/백업_v316_평가표합치기전.xlsx 에 있다.
    'scoring': {
        'general_per_case': 1,
        'S2': {'label': '★', 'name': '중대운영', 'first': 8, 'more': 2, 'cap': 45},
        'S1': {'label': '★★', 'name': '즉시위해', 'first': 12, 'more': 4, 'cap': 48},
        'qsc_floor': 0,
        'final_floor': 0,
    },
    'final_weights': {'qsc_inspector': 0.6, 'mystery_shopper': 0.3, 'improvement': 0.1},
    'grades': [
        {'name': '우수', 'rule': '>= 93'}, {'name': '양호', 'rule': '>= 85'},
        {'name': '보통', 'rule': '>= 76'}, {'name': '미흡', 'rule': '>= 66'},
        {'name': '주의', 'rule': '>= 55'}, {'name': '부적합', 'rule': '< 55'},
    ],
    'texts': {
        'criteria': ws['C2'].value, 'principles': ws['C5'].value,
        'shopper_criteria': sh.cell(2, MSC + 3).value,      # 쇼퍼 C2 = U2
        'shopper_principles': sh.cell(5, MSC + 3).value,    # 쇼퍼 C5 = U5
        'shopper_grade_note': shopper_note,
    },
    'qsc_groups': qsc_groups,
    'shopper_categories': shopper_cats,
}
OUT.parent.mkdir(parents=True, exist_ok=True)

# ★source_sha·version 은 '뽑아낸 내용'에서 낸다 — 엑셀 파일 바이트가 아니다★
#   (2026-08-21) 엑셀을 열었다 그냥 닫아도 파일 바이트와 수정시각이 바뀐다. 그때마다
#   source_sha·version 이 달라져서 master.json이 '바뀐 파일'로 잡히고, 배포 점검이
#   캐시 버전을 올리라고 한다. 문항은 한 글자도 안 바뀌었는데 26곳 폰이 앱을 다시 받는다.
#   그렇게 두 번 헛 배포가 났다(v54·v55). 내용이 같으면 값도 같아야 한다.
#   ★엑셀이 정말 바뀌면 아래 payload 가 달라지므로 여전히 잡힌다.★
_payload = dict(master)
_payload.pop('version', None)
_payload.pop('source_sha', None)
_body = json.dumps(_payload, ensure_ascii=False, sort_keys=True)
master['source_sha'] = hashlib.sha256(_body.encode('utf-8')).hexdigest()[:12]

# version 은 '내용이 마지막으로 바뀐 날'이다 — 내용이 같으면 옛 날짜를 그대로 지킨다.
_prev = {}
if OUT.exists():
    try:
        _prev = json.loads(OUT.read_text(encoding='utf-8'))
    except Exception:
        _prev = {}
if _prev.get('source_sha') == master['source_sha'] and _prev.get('version'):
    master['version'] = _prev['version']
else:
    master['version'] = XLSX_VERSION

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(master, f, ensure_ascii=False, indent=1)

print('QSC', item_no, '문항 /', len(qsc_groups), '그룹 / ★★', sev_count['S1'], '· ★', sev_count['S2'])
print('쇼퍼', q_no, '문항 /', len(shopper_cats), '카테고리 / 관찰', len(YN_ROWS),
      '· 5점 척도', n_likert, '(행 %d~%d)' % (Q_MIN, Q_MAX))
print('저장:', OUT)
